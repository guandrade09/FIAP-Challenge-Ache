import pandas as pd
from extensions import db
from models import Projeto  # certifique-se de que o import está correto
from config import UPLOAD_PATH
import os
import re
import pdfplumber
from openpyxl import load_workbook
import zipfile
import xml.etree.ElementTree as ET

#VERIFICAR COMPATIBILIDADE DOS CABEÇALHO DOS ARQUIVOS
def compatibilidade_arquivo(arquivo):
    filename = arquivo.filename
    caminho_modelo = os.path.join(UPLOAD_PATH, 'Modelo.xlsx')
    df_modelo = pd.read_excel(caminho_modelo)
    try:
        if filename.endswith('.xlsx'):
            df_arquivo = pd.read_excel(arquivo)
        elif filename.endswith('.csv'):
            df_arquivo = pd.read_csv(arquivo)
        else:
            return f'Formato de arquivo inválido'
    
        if list(df_modelo.columns) != list(df_arquivo.columns):
            return f'Arquivo incompativel, cabeçalho invalido!'
        else:
            return True

    except Exception as e:
        return f'Erro na leitura do arquivo!'

#FILTRAR PARA ENCAMINHAR P/ O BD
def filtrar_dados_projetos(caminho_arquivo):
    df = pd.read_excel(caminho_arquivo)
    df = df.dropna(how='all').fillna('')

    df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

    df = df.rename(columns={
        'número': 'numero',
        'classificação': 'classificacao',
        'condição': 'condicao',
        'duração': 'duracao',
        'como_fazer': 'comofazer',
        'documento_referência': 'docreferencia',
        '%_concluída': 'porcentagemconcluida',
    })

    # 🔍 Pega todos os numeros existentes no banco
    numeros_existentes = {p.numero for p in Projeto.query.with_entities(Projeto.numero).all()}

    # 🧹 Remove do DataFrame todas as linhas que já existem no banco
    df_filtrado = df[~df['numero'].isin(numeros_existentes)]

    if df_filtrado.empty:
        raise ValueError("⚠️ Todos os projetos do arquivo já existem no banco.")  # Nada novo para inserir

    return df_filtrado

#UPLOAD
def inserir_dados_projetos(df_filtrado, nome_arquivo):
    try:
        # 1️⃣ Buscar todos os projetos já existentes
        projetos_existentes = Projeto.query.all()
        projetos_existentes_set = {
            (
                str(p.numero).strip().lower(),
                str(p.classificacao).strip().lower(),
                str(p.categoria).strip().lower(),
                str(p.fase).strip().lower(),
                str(p.condicao).strip().lower(),
                str(p.nome).strip().lower(),
                str(p.duracao).strip().lower(),
                str(p.comoFazer).strip().lower(),
                str(p.docReferencia).strip().lower(),
                int(p.porcentagemConcluida)
            )
            for p in projetos_existentes
        }

        novos_projetos = 0
        qnt_duplicata = 0

        # 2️⃣ Percorrer cada linha do DataFrame
        for _, row in df_filtrado.iterrows():
            linha_tupla = (
                str(row['numero']).strip().lower(),
                str(row['classificacao']).strip().lower(),
                str(row['categoria']).strip().lower(),
                str(row['fase']).strip().lower(),
                str(row['condicao']).strip().lower(),
                str(row['nome']).strip().lower(),
                str(row['duracao']).strip().lower(),
                str(row['comofazer']).strip().lower(),
                str(row['docreferencia']).strip().lower(),
                limpar_porcentagem(row['porcentagemconcluida'])
            )

            # 3️⃣ Detectar duplicatas, mas ainda salvar
            if linha_tupla in projetos_existentes_set:
                qnt_duplicata += 1
            else:
                novos_projetos += 1
                projetos_existentes_set.add(linha_tupla)  # adiciona pra evitar contagem dupla

            # 4️⃣ Criar e adicionar ao banco
            projeto = Projeto(
                numero=row['numero'],
                classificacao=row['classificacao'],
                categoria=row['categoria'],
                fase=row['fase'],
                condicao=row['condicao'],
                nome=row['nome'],
                duracao=row['duracao'],
                comoFazer=row['comofazer'],
                docReferencia=row['docreferencia'],
                porcentagemConcluida=limpar_porcentagem(row['porcentagemconcluida']),
                nomeArquivo=nome_arquivo
            )
            db.session.add(projeto)


        total = novos_projetos + qnt_duplicata
        if total == qnt_duplicata:
            return False
        
        print(f"✅ {total} projetos inseridos (novos: {novos_projetos}, duplicados: {qnt_duplicata}).")
        # 5️⃣ Confirmar alterações no banco
        db.session.commit()

        return True

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao inserir projetos: {e}")
        return False

#TRATAMENTO
def tratar_dados_do_arquivo(caminho_arquivo):
    # Lê o arquivo
    if caminho_arquivo.endswith(".xlsx"):
        df = pd.read_excel(caminho_arquivo)
    elif caminho_arquivo.endswith(".csv"):
        df = pd.read_csv(caminho_arquivo)
    else:
        raise ValueError("Formato de arquivo não suportado.")
    
    # Remove linhas totalmente vazias e substitui NaN por string vazia
    df = df.dropna(how='all').fillna('')

    # Padroniza nomes das colunas
    df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

    # Verifica se 'numero' existe após padronização
    if 'número' not in df.columns:
        raise KeyError(f"O arquivo '{caminho_arquivo}' não contém a coluna obrigatória 'numero'. "
                       f"Colunas encontradas: {list(df.columns)}")

    # Renomeia para manter o padrão interno do sistema
    df = df.rename(columns={
        'numero': 'número',
        'classificacao': 'classificação',
        'categoria': 'categoria',
        'fase': 'fase',
        'condicao': 'condição',
        'nome': 'nome',
        'duracao': 'duracao',
        'como_fazer': 'como_fazer',
        'documento_referencia': 'documento_referência',
        '%_concluida': '%_concluída',
    })

    # Limpa porcentagens se a coluna existir
    if '%_concluída' in df.columns:
        df['%_concluída'] = df['%_concluída'].apply(limpar_porcentagem)

    return df

#LIMPAR PORCETAGEM DA COLUNA %_Concluida  !FINALIZADO!
def limpar_porcentagem(valor):
    """
    Normaliza vários formatos de entrada para um inteiro entre 0 e 100.
    Aceita: "95%", "95,0", "95.0", "0.95", 95, 0.95, etc.
    Retorna um int (0..100).
    """
    try:
        if pd.isna(valor) or valor == '':
            return 0

        if isinstance(valor, str):
            s = valor.strip().replace('%', '').replace(' ', '').replace(',', '.')
            if s == '':
                return 0
            f = float(s)
        elif isinstance(valor, (int, float)):
            f = float(valor)
        else:
            return 0

        # Se estiver em 0..1 -> trata como fração (0.95 -> 95)
        if 0 <= f <= 1:
            inteiro = int(round(f * 100))
        else:
            inteiro = int(round(f))

        # limita 0..100
        inteiro = max(0, min(100, inteiro))
        return inteiro

    except Exception:
        return 0

def cruzarDadosPDF(objeto_procurado, PDF_PATH):
    resultados = []

    # padrão genérico para encontrar quaisquer marcadores "TextoX" (com variações)
    padrao_geral_marker = re.compile(r"texto\.?\s*\d+\.?:", flags=re.IGNORECASE)

    # normalizar entrada: extrair número se existir (ex: "Texto.4" -> "4")
    num_match = re.search(r"\d+", str(objeto_procurado))
    numero_procurado = num_match.group() if num_match else None

    # se não achou número, tenta usar o próprio texto normalizado (fallback)
    termo_procurado_raw = str(objeto_procurado).strip()

    for raiz, _, arquivos in os.walk(PDF_PATH):
        for arquivo in arquivos:
            if not arquivo.lower().endswith(".pdf"):
                continue

            caminho_pdf = os.path.join(raiz, arquivo)
            try:
                with pdfplumber.open(caminho_pdf) as pdf:
                    for pagina in pdf.pages:
                        texto_raw = pagina.extract_text() or ""
                        # colapsa espaços e quebras para evitar quebras estranhas atrapalhem o regex
                        texto = re.sub(r"\s+", " ", texto_raw).strip()
                        texto_lower = texto.lower()

                        # primeiro, tenta encontrar pelo número (mais confiável)
                        match_item = None
                        if numero_procurado:
                            # aceita variações: Texto4:, Texto4.:, Texto.4:, Texto.4.:, Texto 4:
                            pattern_num = re.compile(rf"texto\.?\s*{re.escape(numero_procurado)}\.?:", flags=re.IGNORECASE)
                            match_item = pattern_num.search(texto_lower)

                        # fallback: procura o termo literal (normalizado, sem diferenciação de caixa)
                        if not match_item and termo_procurado_raw:
                            # cria termo normalizado (remove múltiplos espaços)
                            termo_norm = re.sub(r"\s+", " ", termo_procurado_raw).lower()
                            # procura a primeira ocorrência do termo no texto
                            idx = texto_lower.find(termo_norm)
                            if idx != -1:
                                # tenta ver se há um marcador completo a partir desse índice (procura "texto...number")
                                # se não for marcador, usaremos idx como início
                                match_item = re.search(r"texto\.?\s*\d+\.?:", texto_lower[idx:]) or None
                                if match_item:
                                    # rebase para índice absoluto
                                    match_item = re.Match  # dummy to be replaced below
                                    # find the absolute match starting at idx
                                    m2 = re.search(r"texto\.?\s*\d+\.?:", texto_lower)
                                    if m2:
                                        match_item = m2
                                    else:
                                        match_item = None
                                else:
                                    # usamos idx como início puro
                                    start_pos = idx
                                    # buscar próximo marcador a partir de start_pos
                                    next_m = padrao_geral_marker.search(texto_lower[start_pos + 1:])
                                    if next_m:
                                        fim = start_pos + 1 + next_m.start()
                                    else:
                                        fim = len(texto)
                                    trecho = texto[start_pos:fim].strip()
                                    resultados.append({
                                        "arquivo": arquivo,
                                        "pagina": pagina.page_number,
                                        "trecho": trecho,
                                        "caminho": caminho_pdf
                                    })
                                    # pular para próxima página
                                    continue

                        # se encontrou o marcador do item por regex
                        if match_item:
                            # se match_item é um objeto re.Match, obter posições absolutas
                            if hasattr(match_item, "start"):
                                start_abs = match_item.start()
                                # encontra próximo marcador depois do atual para delimitar até onde vai
                                next_search = padrao_geral_marker.search(texto_lower[start_abs + match_item.end() - match_item.start():])
                                if next_search:
                                    fim_abs = start_abs + (match_item.end() - match_item.start()) + next_search.start()
                                else:
                                    fim_abs = len(texto)
                                trecho = texto[start_abs:fim_abs].strip()
                                resultados.append({
                                    "arquivo": arquivo,
                                    "pagina": pagina.page_number,
                                    "trecho": trecho,
                                    "caminho": caminho_pdf
                                })
                            else:
                                # segurança: se match_item não for match object, ignora
                                pass
                        # else: nada encontrado nesta página (continua)
            except Exception as e:
                print(f"Erro abrindo {caminho_pdf}: {e}")

    return resultados

def extrair_hyperlinks(caminho, nome_coluna="documento_referência"):
    wb = load_workbook(caminho, data_only=False)
    ws = wb.active

    # Encontrar a coluna certa
    col_index = None
    for cell in ws[1]:
        if str(cell.value).strip().lower() == nome_coluna.lower():
            col_index = cell.column
            break

    if not col_index:
        return []

    links = []

    # Percorrer todas relações internas da planilha
    rels = ws._rels

    for row in ws.iter_rows(min_row=2):
        cell = row[col_index - 1]

        # 1) Caso fosse hyperlink padrão
        if cell.hyperlink:
            links.append({"texto": cell.value, "url": cell.hyperlink.target})
            continue

        # 2) Caso seja hyperlink "oculto" do SharePoint (seu caso)
        # cada célula pode ter id -> relação
        if cell._value is not None and hasattr(cell, '_hyperlink'):
            rel_id = cell._hyperlink.id
            if rel_id in rels:
                url = rels[rel_id].target
                links.append({"texto": cell.value, "url": url})
                continue

        # 3) Caso link esteja salvo como Rich-Text (estilo Word-like)
        if hasattr(cell, 'value') and hasattr(cell.value, 'text'):
            if hasattr(cell.value, 'hyperlink') and cell.value.hyperlink:
                links.append({"texto": cell.value.text, "url": cell.value.hyperlink})
                continue

    return links

def extrair_hyperlinks_sharepoint(caminho, nome_coluna="documento_referência"):
    wb = load_workbook(caminho, data_only=False)
    ws = wb.active

    # Descobrir índice da coluna
    col_index = None
    for cell in ws[1]:
        if str(cell.value).strip().lower() == nome_coluna.lower():
            col_index = cell.column
            break

    if not col_index:
        return []

    # Extrair sharedStrings.xml (onde ficam hyperlinks modernos do SharePoint)
    hyperlinks = []
    with zipfile.ZipFile(caminho, 'r') as z:
        if 'xl/sharedStrings.xml' in z.namelist():
            xml = z.read('xl/sharedStrings.xml')
            root = ET.fromstring(xml)

            # Namespaces
            ns = {'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

            shared_strings = []
            for si in root.findall('a:si', ns):
                text = ''.join(t.text or '' for t in si.findall('.//a:t', ns))
                hyperlink = si.find('.//a:rPh', ns)
                shared_strings.append(text)

    # Agora percorre a planilha e verifica se o texto bate com sharedStrings
    for row in ws.iter_rows(min_row=2):
        cell = row[col_index - 1]

        if cell.value and isinstance(cell.value, str):
            texto = cell.value.strip()
            # Procurar URL associada ao texto no sharedStrings
            for s in shared_strings:
                if texto in s and "http" in s:
                    url = s[s.index("http"):]  # pega parte da URL em diante
                    hyperlinks.append({"texto": texto, "url": url})
                    break

    return hyperlinks